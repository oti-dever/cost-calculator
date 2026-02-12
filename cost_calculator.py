"""
Excel文件处理脚本 - 成本计算与店铺统计
================================================================================

功能概述：
本脚本用于处理电商订单Excel文件，自动计算各类成本并按店铺汇总统计。
输入原始Excel文件，输出包含成本明细和店铺统计的新Excel文件。

核心功能：
1. 成本计算
   - 基础成本：根据size_material_price.json中的尺寸、材质、关键词匹配计算
   - 代发成本：识别"枕芯"、"yr"/"义乳"、"yt"/"义臀"等关键词，每个加5.5元
   - 义乳/义臀成本：从moving_and_selling_costs.json匹配完整remark计算
   - 枕芯成本：从pillow_cost.json匹配尺寸+枕芯类型计算
   - 硅胶/电动成本：从others.json匹配关键词计算

2. 店铺识别
   - 自动从"商家/店铺"列提取店铺名称（去空白、规范化处理）
   - 不再依赖shop.json配置文件

3. 输出结构（两个Sheet）
   a) 成本明细Sheet
      - 保留原始数据的所有列
      - 新增列：成本、代发成本、总成本（=成本+代发成本）
      - 新增列（空列后）：义乳/义臀成本、枕芯成本、硅胶/电动成本
      - 底部添加合计行，使用SUM公式汇总各成本列

   b) 店铺统计Sheet
      - 按店铺汇总各类成本（成本、代发成本、义乳/义臀、枕芯、硅胶/电动、总成本）
      - 使用SUMIF公式动态引用成本明细Sheet数据
      - 支持手动修改成本明细后自动更新统计
      - 底部添加合计行

4. 特殊处理
   - 海外订单：包含"发海外"关键词的订单，所有成本列置空，需手动补全
   - 数量识别：通过统计逗号（半角/全角）数量来判断订单项数量
   - 错误标记：义乳/义臀或枕芯成本匹配失败时，对应列置空

输入要求：
- Excel文件必须包含"卖家备注"和"商家/店铺"列
- 支持.xlsx和.xls格式（.xls可能无法保留原始样式）

配置文件（位于脚本同目录）：
- size_material_price.json：基础价格配置
- moving_and_selling_costs.json：义乳/义臀成本配置
- pillow_cost.json：枕芯成本配置
- others.json：硅胶/电动成本配置

输出文件：
- 文件名：${原文件名}-已处理.xlsx
- 位置：与原文件相同目录

使用方法：
1. 命令行运行：python price_calculator.py <Excel文件路径>
2. 交互式运行：python price_calculator.py（然后输入文件路径）
3. 打包为exe后拖拽Excel文件到exe图标上运行

版本信息：
- 支持多Sheet输出（成本明细 + 店铺统计）
- 支持多种成本类型计算
- 使用Excel公式实现动态更新
================================================================================
"""

import os
import json
import re
import sys
import glob
import shlex
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Sequence, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd


def resolve_resource_path(filename: str) -> str:
    """解析资源文件路径。

    规则：
    1. 打包运行时（PyInstaller onefile）：
       - 优先读取 exe 同目录下的外部文件（便于紧急覆盖配置）
       - 若外部不存在，则读取 exe 内嵌资源（_MEIPASS）
    2. 脚本运行时：读取脚本同目录文件

    参数:
        filename (str): 资源文件名

    返回:
        str: 可访问的资源路径（若均不存在则返回首选候选路径）
    """

    if getattr(sys, "frozen", False):
        exe_dir = os.path.dirname(sys.executable)
        external_path = os.path.join(exe_dir, filename)
        if os.path.exists(external_path):
            return external_path

        bundle_dir = getattr(sys, "_MEIPASS", exe_dir)
        bundled_path = os.path.join(bundle_dir, filename)
        if os.path.exists(bundled_path):
            return bundled_path

        return external_path

    script_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(script_dir, filename)


@dataclass
class ProcessResult:
    """单个文件处理结果。"""

    input_path: str
    success: bool
    status: str
    message: str
    output_path: Optional[str] = None
    matched_count: Optional[int] = None
    total_count: Optional[int] = None
    overseas_count: int = 0


def build_output_file_path(
    input_file_path: str,
    output_dir: Optional[str] = None,
    overwrite: bool = False,
) -> str:
    """构建输出文件路径。

    参数:
        input_file_path (str): 输入文件路径
        output_dir (Optional[str]): 输出目录；为空时默认输出到输入文件目录
        overwrite (bool): 是否允许覆盖同名文件

    返回:
        str: 最终输出文件路径
    """

    source_dir = os.path.dirname(input_file_path)
    target_dir = output_dir if output_dir else source_dir
    file_name = os.path.basename(input_file_path)
    name_without_ext, _ext = os.path.splitext(file_name)
    base_name = f"{name_without_ext}-已处理"
    output_path = os.path.join(target_dir, f"{base_name}.xlsx")

    if overwrite or not os.path.exists(output_path):
        return output_path

    index = 1
    while True:
        candidate = os.path.join(target_dir, f"{base_name}({index}).xlsx")
        if not os.path.exists(candidate):
            return candidate
        index += 1


def create_session_log(base_path: str) -> Tuple[Optional[str], Callable[[str], None]]:
    """创建会话日志文件，并返回日志函数。"""

    try:
        logs_dir = os.path.join(base_path, "logs")
        os.makedirs(logs_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = os.path.join(logs_dir, f"cost_calculator_{timestamp}.log")

        def _write_log(message: str) -> None:
            time_prefix = datetime.now().strftime("%H:%M:%S")
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"[{time_prefix}] {message}\n")

        _write_log("日志已启动")
        return log_path, _write_log
    except Exception:
        return None, lambda _msg: None


def print_batch_summary(results: List[ProcessResult], write_log: Callable[[str], None]) -> None:
    """打印批处理汇总与失败清单。"""

    if not results:
        print("本批次没有可处理项。")
        write_log("本批次没有可处理项")
        return

    success_items = [r for r in results if r.status == "success"]
    failed_items = [r for r in results if r.status == "failed"]
    skipped_items = [r for r in results if r.status == "skipped"]

    total_files = len(results)
    total_overseas = sum(r.overseas_count for r in success_items)

    summary_line = (
        f"批处理汇总: 总计 {total_files} | 成功 {len(success_items)} | "
        f"失败 {len(failed_items)} | 跳过 {len(skipped_items)}"
    )
    print(summary_line)
    write_log(summary_line)

    if total_overseas > 0:
        overseas_line = f"海外订单累计: {total_overseas} 条"
        print(overseas_line)
        write_log(overseas_line)

    if failed_items:
        print("失败清单:")
        write_log("失败清单:")
        for idx, item in enumerate(failed_items, start=1):
            line = f"  {idx}. {item.input_path} | 原因: {item.message}"
            print(line)
            write_log(line)

    if skipped_items:
        print("跳过清单:")
        write_log("跳过清单:")
        for idx, item in enumerate(skipped_items, start=1):
            line = f"  {idx}. {item.input_path} | 原因: {item.message}"
            print(line)
            write_log(line)


def load_price_data(json_path):
    """加载并根据优先级排序价格数据"""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
        # 根据 "priority" 字段对类别进行排序
        return sorted(data, key=lambda x: x.get("priority", 99))


def load_shop_data(json_path):
    """兼容旧接口：不再使用 shop.json，始终返回空字典"""
    print("提示: 店铺统计改为自动整理‘商家/店铺’列，不再读取 shop.json。")
    return {}


def load_moving_costs(json_path):
    """加载动销成本数据"""
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"警告: 找不到动销成本配置文件 {json_path}，将跳过动销成本功能")
        return []


def load_pillow_cost(json_path):
    """加载枕芯成本数据"""
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"警告: 找不到枕芯成本配置文件 {json_path}，将跳过枕芯成本功能")
        return {}


def load_others_cost(json_path):
    """加载硅胶/电动成本数据"""
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"警告: 找不到硅胶/电动成本配置文件 {json_path}，将跳过硅胶/电动成本功能")
        return []


def identify_shop(text, shop_data):
    """
    识别文本中的店铺名称（从商家/店铺列进行包含匹配，左起最先命中）。

    规则：
    - 仅使用店铺名称进行识别（不使用编号解析）。
    - 规范化：移除空白（含全角）、转为小写，以提升兼容度。
    - 匹配方式：从左到右，返回最先出现的店铺名称。

    参数:
        text (str): 商家/店铺列的文本
        shop_data (dict): 店铺编号到店铺名称的映射

    返回:
        str: 店铺名称，如果未找到则返回 None
    """
    if pd.isna(text) or not text:
        return None

    # 规范化：去空白（半角/全角）、小写
    text_str = re.sub(r"\s+|\u3000", "", str(text)).lower()

    # 构造候选名称列表
    name_candidates = list(shop_data.values())

    # 使用从左到右的最先匹配策略
    matched, _pos = find_leftmost_match(text_str, name_candidates)
    if matched is not None:
        return matched

    return None


def find_leftmost_match(text, candidates):
    """
    从左向右找第一个（最左边）匹配的候选项

    参数:
        text (str): 要搜索的文本（已转为小写）
        candidates (list): 候选项列表（字符串或元组）

    返回:
        tuple: (匹配的候选项, 位置)，如果没有匹配则返回 (None, float('inf'))
    """
    best_match = None
    best_pos = float("inf")

    for candidate in candidates:
        # 如果候选项是元组，取第一个元素作为搜索字符串
        search_str = candidate[0] if isinstance(candidate, tuple) else candidate
        pos = text.find(search_str.lower())
        if pos != -1 and pos < best_pos:
            best_match = candidate
            best_pos = pos

    return best_match, best_pos


def find_longest_match_at_leftmost(text, candidates):
    """
    从左向右找匹配,如果多个候选项在同一位置匹配,选择最长的
    专用于材质匹配,按长度从长到短优先

    参数:
        text (str): 要搜索的文本（已转为小写）
        candidates (list): 候选项列表（已按长度从长到短排序）

    返回:
        tuple: (匹配的候选项, 位置)，如果没有匹配则返回 (None, float('inf'))
    """
    best_match = None
    best_pos = float("inf")
    best_len = 0

    for candidate in candidates:
        # 如果候选项是元组，取第一个元素作为搜索字符串
        search_str = candidate[0] if isinstance(candidate, tuple) else candidate
        pos = text.find(search_str.lower())
        if pos != -1:
            # 优先级: 1. 位置更靠左 2. 同一位置时长度更长
            if pos < best_pos or (pos == best_pos and len(search_str) > best_len):
                best_match = candidate
                best_pos = pos
                best_len = len(search_str)

    return best_match, best_pos


def match_price(
    text,
    price_data,
    moving_costs_data=None,
    pillow_cost_data=None,
    others_cost_data=None,
):
    """
    根据文本内容（卖家备注）匹配价格并计算所有成本。

    一条记录的处理逻辑：
    1. 将输入文本按分隔符（'。'）分割成多个独立的订单项。
    2. 对每个订单项：
       a. 在 `price_data` 中按优先级（priority）查找匹配的关键词（keywords）、尺寸和材质，以确定基础价格。
       b. 使用正则表达式解析订单项中的数量（例如订单编号 "B-5634"）。
       c. 调用单项计算函数计算: 义乳/义臀成本、枕芯成本、硅胶/电动成本
       d. 计算该订单项的成本（基础价格 * 数量）和代发成本（附加费用）。
    3. 将所有订单项的价格累加，返回各类成本。

    参数:
        text (str): 要匹配的文本。
        price_data (list): 从JSON加载的价格数据，每项包含 keywords、priority、products 等字段。
        moving_costs_data (list): 义乳/义臀成本数据
        pillow_cost_data (dict): 枕芯成本数据
        others_cost_data (list): 硅胶/电动成本数据

    返回:
        tuple: (成本, 代发成本, 总成本, 义乳/义臀成本, 枕芯成本, 硅胶/电动成本)
               如果未找到任何匹配项，则返回 (None, None, None, 0, 0, 0)
               如果义乳/义臀或枕芯有匹配错误，对应项返回 None
    """
    if pd.isna(text) or not text:
        return None, None, None, None, None, None

    # 去除文本中所有空格
    text = re.sub(
        r"\s+", "", str(text)
    )  # 去除所有空白字符（包括半角空格、全角空格、制表符等）

    # 去除半角和全角空格，并使用'。'分割项目
    items = str(text).replace(" ", "").replace("\u3000", "").split("。")
    total_base_cost = 0  # 成本（基础价格）
    total_dropship_cost = 0  # 代发成本（附加费用）
    total_moving_cost = 0.0  # 义乳/义臀成本
    total_pillow_cost = 0.0  # 枕芯成本
    total_others_cost = 0.0  # 硅胶/电动成本

    # 用于标记是否有匹配错误
    moving_has_error = False
    pillow_has_error = False

    for item_text in items:
        if not item_text.strip():
            continue

        # 转换为小写以便忽略大小写匹配
        item_text_lower = item_text.lower()
        item_price = 0
        base_price_found = False

        # 检查数量：统计半角逗号和全角逗号的数量
        quantity = item_text.count(",") + item_text.count("，")

        # 查找基础价格 - 按优先级排序后匹配
        sorted_categories = sorted(price_data, key=lambda x: x.get("priority", 999))

        for category_data in sorted_categories:
            keywords = category_data.get("keywords", [])
            products = category_data["products"]

            # 检查是否有任何关键词匹配（从左向右找第一个匹配的）
            matched_keyword, keyword_pos = find_leftmost_match(
                item_text_lower, keywords
            )
            if matched_keyword is None:
                continue

            # 从左向右找第一个匹配的尺寸
            # 为每个产品准备尺寸候选项（包含多种格式）
            size_candidates = []
            for product in products:
                size = product["尺寸"]
                # 准备三种尺寸格式
                size_formats = [
                    size,  # 50*150
                    size.replace("*", "x"),  # 50x150
                    size.replace("*", ""),  # 50150
                ]
                # 将产品和尺寸格式组合为元组
                for size_format in size_formats:
                    size_candidates.append((size_format, product))

            matched_size, size_pos = find_leftmost_match(
                item_text_lower, size_candidates
            )
            if matched_size is None:
                continue

            # 获取匹配的产品
            matched_product = matched_size[1]

            # 材质匹配: 按材质名称长度从长到短排序,优先匹配更长的材质名
            material_candidates = [
                (material, price) for material, price in matched_product["价格"].items()
            ]
            material_candidates.sort(key=lambda x: len(x[0]), reverse=True)
            matched_material, material_pos = find_longest_match_at_leftmost(
                item_text_lower, material_candidates
            )

            if matched_material:
                item_price = matched_material[1]
                base_price_found = True
                break

        # 处理数量为0的情况
        if quantity == 0:
            if base_price_found:
                # 如果匹配到价格，按数量1计算
                quantity = 1
            # 如果没有匹配到价格，quantity保持为0，只统计额外费用

        # 如果没有匹配到价格且数量>0，则返回空
        if not base_price_found and quantity > 0:
            return None, None, None, None, None, None

        # 计算成本（基础价格 * 数量）
        base_cost = item_price * quantity
        total_base_cost += base_cost

        # 检查额外费用（代发成本），对关键词进行计数
        extra_cost = 0

        # 统计"枕芯"出现的次数
        pillow_count = item_text_lower.count("枕芯")
        extra_cost += pillow_count * 5.5

        # 统计"yr"或"义乳"出现的次数
        yr_count = item_text_lower.count("yr") + item_text_lower.count("义乳")
        extra_cost += yr_count * 5.5

        # 统计"yt"或"义臀"出现的次数
        yt_count = item_text_lower.count("yt") + item_text_lower.count("义臀")
        extra_cost += yt_count * 5.5

        # 代发成本 = 附加费用（不乘以数量）
        dropship_cost = extra_cost
        total_dropship_cost += dropship_cost

        # 在遍历每个订单项时,调用单项计算函数
        # 计算义乳/义臀成本
        if moving_costs_data and not moving_has_error:
            item_moving_cost = calculate_moving_cost_for_item(
                item_text.strip(), moving_costs_data
            )
            if item_moving_cost is None:
                moving_has_error = True
            else:
                total_moving_cost += item_moving_cost

        # 计算枕芯成本
        if pillow_cost_data and not pillow_has_error:
            item_pillow_cost = calculate_pillow_cost_for_item(
                item_text.strip(), pillow_cost_data
            )
            if item_pillow_cost is None:
                pillow_has_error = True
            else:
                total_pillow_cost += item_pillow_cost

        # 计算硅胶/电动成本
        if others_cost_data:
            item_others_cost = calculate_others_cost_for_item(
                item_text.strip(), others_cost_data
            )
            total_others_cost += item_others_cost

        # 如果数量为0，没有匹配到价格，且没有额外费用，返回空
        if (
            quantity == 0
            and not base_price_found
            and extra_cost == 0
            and total_others_cost == 0
        ):
            return None, None, None, None, None, None

    total_cost = total_base_cost + total_dropship_cost

    # 如果有匹配错误,返回None
    final_moving_cost = None if moving_has_error else total_moving_cost
    final_pillow_cost = None if pillow_has_error else total_pillow_cost

    return (
        total_base_cost,
        total_dropship_cost,
        total_cost,
        final_moving_cost,
        final_pillow_cost,
        total_others_cost,
    )


def calculate_moving_cost_for_item(item_text, moving_costs_data):
    """
    为单个订单项计算义乳/义臀成本（严格匹配）

    规则：
    - 如果订单项不包含yr或yt关键词，返回0
    - 如果包含yr或yt，必须每个yr/yt都能在moving_and_selling_costs.json中匹配到完整remark
    - 搜索范围: 上一个yr/yt位置的末尾后(如有)到下一个yr/yt位置的开头前(如有)
    - 如果任何一个yr/yt匹配不到，返回None（表示备注错误）

    参数:
        item_text (str): 单个订单项文本（已去除空格）
        moving_costs_data (list): 动销成本数据列表

    返回:
        float or None: 该订单项的义乳/义臀成本，如果有匹配错误返回 None，无yr/yt关键词返回 0
    """
    if not item_text or not moving_costs_data:
        return 0

    item_lower = item_text.lower()

    # 检查是否包含yr或yt关键词
    if "yr" not in item_lower and "yt" not in item_lower:
        return 0

    # 构建remark字典（小写）
    remark_dict = {}
    for item in moving_costs_data:
        remark = str(item.get("remark", "")).lower()
        if remark:
            remark_dict[remark] = float(item.get("cost", 0))

    # 找出所有yr和yt的位置(起始位置+长度)
    positions = []  # [(start_pos, end_pos, keyword)]

    # 查找yr
    pos = 0
    while True:
        pos = item_lower.find("yr", pos)
        if pos == -1:
            break
        positions.append((pos, pos + 2, "yr"))
        pos += 1

    # 查找yt
    pos = 0
    while True:
        pos = item_lower.find("yt", pos)
        if pos == -1:
            break
        positions.append((pos, pos + 2, "yt"))
        pos += 1

    if not positions:
        return 0

    # 按起始位置排序
    positions.sort(key=lambda x: x[0])

    item_cost = 0.0

    # 对每个yr/yt位置，在限定范围内尝试匹配完整的remark
    for i, (start_pos, end_pos, keyword) in enumerate(positions):
        matched = False

        # 确定搜索范围: 上一个yr/yt末尾后 到 下一个yr/yt开头前
        search_start = positions[i - 1][1] if i > 0 else 0  # 上一个的末尾
        search_end = (
            positions[i + 1][0] if i < len(positions) - 1 else len(item_lower)
        )  # 下一个的开头
        search_text = item_lower[search_start:search_end]

        # 按remark长度从长到短排序，优先匹配更长的
        sorted_remarks = sorted(
            remark_dict.items(), key=lambda x: len(x[0]), reverse=True
        )

        for remark, cost in sorted_remarks:
            # 检查remark是否包含当前关键词
            if keyword not in remark:
                continue

            # 在搜索范围内查找remark
            if remark in search_text:
                item_cost += cost
                matched = True
                break

        # 如果这个yr/yt没有匹配到任何remark，返回None表示错误
        if not matched:
            return None

    return item_cost


def calculate_pillow_cost_for_item(item_text, pillow_cost_data):
    """
    为单个订单项计算枕芯成本（严格匹配）

    规则：
    - 如果订单项不包含"枕芯"关键词，返回0
    - 如果包含"枕芯"，必须每个枕芯都能匹配到 尺寸+枕芯类型关键词
    - 搜索范围: 上一个枕芯位置后(如有)到当前枕芯位置
    - 使用正则表达式匹配: {尺寸}{0-n个空格}{枕芯关键词}
    - 如果任何一个枕芯匹配不到，返回None（表示备注错误）

    参数:
        item_text (str): 单个订单项文本（已去除空格）
        pillow_cost_data (dict): 枕芯成本数据字典 {size: {keyword: cost}}

    返回:
        float or None: 该订单项的枕芯成本，如果有匹配错误返回 None，无枕芯关键词返回 0
    """
    if not item_text or not pillow_cost_data:
        return 0

    item_lower = item_text.lower()

    # 检查是否包含"枕芯"关键词
    if "枕芯" not in item_lower:
        return 0

    # 找出所有"枕芯"的位置
    pillow_positions = []
    pos = 0
    while True:
        pos = item_lower.find("枕芯", pos)
        if pos == -1:
            break
        pillow_positions.append(pos)
        pos += 2  # "枕芯"是2个字符，跳过以避免重复

    if not pillow_positions:
        return 0

    item_cost = 0.0

    # 对每个"枕芯"位置，在限定范围内尝试匹配尺寸和关键词
    for i, pillow_pos in enumerate(pillow_positions):
        matched = False

        # 确定搜索范围: 上一个枕芯位置后 到 当前枕芯位置(包含)
        search_start = pillow_positions[i - 1] + 2 if i > 0 else 0  # 上一个枕芯的末尾
        search_end = (
            pillow_positions[i + 1] + 2
            if i < len(pillow_positions) - 1
            else len(item_lower)
        )  # 当前枕芯的末尾
        search_text = item_lower[search_start:search_end]

        # 按尺寸从长到短排序
        sorted_sizes = sorted(pillow_cost_data.keys(), key=len, reverse=True)

        for size in sorted_sizes:
            size_lower = size.lower()
            # 准备三种尺寸格式
            size_formats = [
                size_lower,  # 120*40
                size_lower.replace("*", "x"),  # 120x40
                size_lower.replace("*", ""),  # 12040
                size_lower.split("*")[1] + "*" + size_lower.split("*")[0],  # 40*120
                size_lower.split("*")[1] + "x" + size_lower.split("*")[0],  # 40x120
                size_lower.split("*")[1] + size_lower.split("*")[0],  # 40120
            ]

            # 获取该尺寸下的所有关键词
            keywords = pillow_cost_data[size]
            sorted_keywords = sorted(keywords.keys(), key=len, reverse=True)

            # 尝试匹配每种尺寸格式 + 每个关键词的组合
            for size_format in size_formats:
                for keyword in sorted_keywords:
                    keyword_lower = keyword.lower()
                    # 构建正则表达式: 关键词(已包含"枕芯") + 0-n个空格 + 尺寸
                    # 注意：item_text已经去除空格，所以实际上不会有空格，但为了健壮性保留这个逻辑
                    pattern = re.escape(keyword_lower) + r"\s*" + re.escape(size_format)

                    if re.search(pattern, search_text):
                        item_cost += keywords[keyword]
                        matched = True
                        break

                if matched:
                    break

            if matched:
                break

        # 如果这个"枕芯"没有匹配到，返回None表示错误
        if not matched:
            return None

    return item_cost


def calculate_others_cost_for_item(item_text, others_cost_data):
    """
    为单个订单项计算硅胶/电动成本

    规则：
    - 匹配others.json中的remark关键词
    - 按最长匹配原则

    参数:
        item_text (str): 单个订单项文本（已去除空格）
        others_cost_data (list): 硅胶/电动成本数据列表

    返回:
        float: 该订单项的硅胶/电动成本，如果没有匹配则返回 0
    """
    if not item_text or not others_cost_data:
        return 0

    item_lower = item_text.lower()

    # 构建remark字典（小写）
    remark_dict = {}
    for item in others_cost_data:
        remark = str(item.get("remark", "")).lower()
        if remark:
            remark_dict[remark] = float(item.get("cost", 0))

    # 按remark长度从长到短排序（优先匹配更长的）
    sorted_remarks = sorted(remark_dict.items(), key=lambda x: len(x[0]), reverse=True)

    item_cost = 0.0

    # 匹配所有可能的remark
    for remark, cost in sorted_remarks:
        if remark in item_lower:
            item_cost += cost
            # 只匹配一次最长的
            break

    return item_cost


def process_shop_summary_sheet(
    workbook,
    detail_sheet_name,
    shop_names,
    data_start_row,
    data_end_row,
    shop_name_col,
    cost_col,
    moving_col,
    dropship_col,
    total_col,
    pillow_col,
    others_col,
):
    """
    创建独立的店铺统计sheet（使用公式引用成本明细sheet）

    参数:
        workbook: openpyxl 工作簿对象
        detail_sheet_name: 成本明细sheet的名称
        shop_names: 店铺名称集合
        data_start_row: 数据起始行（通常是2）
        data_end_row: 数据结束行
        shop_name_col: 店铺名称列的字母
        cost_col: 成本列的字母
        moving_col: 义乳/义臀成本列的字母
        dropship_col: 代发成本列的字母
        total_col: 总成本列的字母
        pillow_col: 枕芯成本列的字母
        others_col: 硅胶/电动成本列的字母
    """
    from openpyxl.styles import Alignment

    # 创建或获取店铺统计sheet
    if "店铺统计" in workbook.sheetnames:
        summary_sheet = workbook["店铺统计"]
        # 清空现有内容
        workbook.remove(summary_sheet)

    summary_sheet = workbook.create_sheet("店铺统计")

    # 写入表头
    header_row = 1
    summary_sheet.cell(row=header_row, column=1, value="店铺名称")
    summary_sheet.cell(row=header_row, column=2, value="成本")
    summary_sheet.cell(row=header_row, column=3, value="代发成本")
    summary_sheet.cell(row=header_row, column=4, value="义乳/义臀成本")
    summary_sheet.cell(row=header_row, column=5, value="枕芯成本")
    summary_sheet.cell(row=header_row, column=6, value="硅胶/电动成本")
    summary_sheet.cell(row=header_row, column=7, value="总成本")

    # 写入店铺数据
    current_row = header_row + 1
    for shop_name in sorted(shop_names):
        # 店铺名称
        summary_sheet.cell(row=current_row, column=1, value=shop_name)

        # 成本合计 - 使用SUMIF公式引用成本明细sheet
        cost_formula = f'=SUMIF({detail_sheet_name}!${shop_name_col}${data_start_row}:${shop_name_col}${data_end_row},"{shop_name}",{detail_sheet_name}!${cost_col}${data_start_row}:${cost_col}${data_end_row})'
        summary_sheet.cell(row=current_row, column=2, value=cost_formula)

        # 代发成本合计
        dropship_formula = f'=SUMIF({detail_sheet_name}!${shop_name_col}${data_start_row}:${shop_name_col}${data_end_row},"{shop_name}",{detail_sheet_name}!${dropship_col}${data_start_row}:${dropship_col}${data_end_row})'
        summary_sheet.cell(row=current_row, column=3, value=dropship_formula)

        # 义乳/义臀成本合计
        moving_formula = f'=SUMIF({detail_sheet_name}!${shop_name_col}${data_start_row}:${shop_name_col}${data_end_row},"{shop_name}",{detail_sheet_name}!${moving_col}${data_start_row}:${moving_col}${data_end_row})'
        summary_sheet.cell(row=current_row, column=4, value=moving_formula)

        # 枕芯成本合计
        pillow_formula = f'=SUMIF({detail_sheet_name}!${shop_name_col}${data_start_row}:${shop_name_col}${data_end_row},"{shop_name}",{detail_sheet_name}!${pillow_col}${data_start_row}:${pillow_col}${data_end_row})'
        summary_sheet.cell(row=current_row, column=5, value=pillow_formula)

        # 硅胶/电动成本合计
        others_formula = f'=SUMIF({detail_sheet_name}!${shop_name_col}${data_start_row}:${shop_name_col}${data_end_row},"{shop_name}",{detail_sheet_name}!${others_col}${data_start_row}:${others_col}${data_end_row})'
        summary_sheet.cell(row=current_row, column=6, value=others_formula)

        # 总成本 = 成本 + 代发成本 + 义乳/义臀成本 + 枕芯成本 + 硅胶/电动成本
        total_row_formula = f"=B{current_row}+C{current_row}+D{current_row}+E{current_row}+F{current_row}"
        summary_sheet.cell(row=current_row, column=7, value=total_row_formula)

        current_row += 1

    # 添加总计行
    summary_sheet.cell(row=current_row, column=1, value="合计")

    # 总计使用SUM公式汇总上面的统计
    sum_start_row = header_row + 1
    sum_end_row = current_row - 1

    # 总计行：分别汇总六个成本列
    summary_sheet.cell(
        row=current_row, column=2, value=f"=SUM(B{sum_start_row}:B{sum_end_row})"
    )
    summary_sheet.cell(
        row=current_row, column=3, value=f"=SUM(C{sum_start_row}:C{sum_end_row})"
    )
    summary_sheet.cell(
        row=current_row, column=4, value=f"=SUM(D{sum_start_row}:D{sum_end_row})"
    )
    summary_sheet.cell(
        row=current_row, column=5, value=f"=SUM(E{sum_start_row}:E{sum_end_row})"
    )
    summary_sheet.cell(
        row=current_row, column=6, value=f"=SUM(F{sum_start_row}:F{sum_end_row})"
    )
    summary_sheet.cell(
        row=current_row, column=7, value=f"=SUM(G{sum_start_row}:G{sum_end_row})"
    )

    print(f"已创建独立的店铺统计sheet（共 {len(shop_names)} 家店铺）")


def process_cost_detail_sheet(
    sheet,
    base_costs,
    dropship_costs,
    moving_costs,
    pillow_costs,
    others_costs,
    shop_names,
    shop_name_col_idx,
):
    """
    处理成本明细sheet,添加成本相关列和合计行

    参数:
        sheet: openpyxl工作表对象
        base_costs: 成本列表
        dropship_costs: 代发成本列表
        moving_costs: 义乳/义臀成本列表
        pillow_costs: 枕芯成本列表
        others_costs: 硅胶/电动成本列表
        shop_names: 店铺名称列表
        shop_name_col_idx: 店铺名称列索引

    返回:
        tuple: (cost_col_idx, dropship_col_idx, grand_total_col_idx, moving_col_idx, pillow_col_idx, others_col_idx)
    """
    # 查找或创建表头
    header_row = sheet[1]
    if header_row is None:
        print("警告: 无法读取表头")
        return None

    header = [cell.value for cell in header_row]

    # 按正确顺序创建列：成本、代发成本、总成本、空列、义乳/义臀成本、枕芯成本、硅胶/电动成本

    # 处理"成本"列
    if "成本" in header:
        cost_col_idx = header.index("成本") + 1
    else:
        cost_col_idx = sheet.max_column + 1
        sheet.cell(row=1, column=cost_col_idx, value="成本")

    # 处理"代发成本"列
    if "代发成本" in header:
        dropship_col_idx = header.index("代发成本") + 1
    else:
        dropship_col_idx = sheet.max_column + 1
        sheet.cell(row=1, column=dropship_col_idx, value="代发成本")

    # 处理"总成本"列（仅成本+代发成本）
    if "总成本" in header:
        grand_total_col_idx = header.index("总成本") + 1
    else:
        grand_total_col_idx = sheet.max_column + 1
        sheet.cell(row=1, column=grand_total_col_idx, value="总成本")

    # 插入一个空列（显示占位，不参与计算）
    empty_col_idx = sheet.max_column + 1
    sheet.cell(row=1, column=empty_col_idx, value="")

    # 处理"义乳/义臀成本"列（在空列之后）
    if "义乳/义臀成本" in header:
        moving_col_idx = header.index("义乳/义臀成本") + 1
    else:
        moving_col_idx = sheet.max_column + 1
        sheet.cell(row=1, column=moving_col_idx, value="义乳/义臀成本")

    # 处理"枕芯成本"列
    if "枕芯成本" in header:
        pillow_col_idx = header.index("枕芯成本") + 1
    else:
        pillow_col_idx = sheet.max_column + 1
        sheet.cell(row=1, column=pillow_col_idx, value="枕芯成本")

    # 处理"硅胶/电动成本"列
    if "硅胶/电动成本" in header:
        others_col_idx = header.index("硅胶/电动成本") + 1
    else:
        others_col_idx = sheet.max_column + 1
        sheet.cell(row=1, column=others_col_idx, value="硅胶/电动成本")

    # 写入数据（列顺序：成本、代发成本、总成本(成本+代发)、空列、义乳/义臀成本、枕芯成本、硅胶/电动成本）
    for i, (
        shop_name,
        base_cost,
        dropship_cost,
        moving_cost,
        pillow_cost,
        others_cost,
    ) in enumerate(
        zip(
            shop_names,
            base_costs,
            dropship_costs,
            moving_costs,
            pillow_costs,
            others_costs,
        )
    ):
        row_index = i + 2
        if shop_name_col_idx:
            sheet.cell(row=row_index, column=shop_name_col_idx, value=shop_name)
        # 成本
        sheet.cell(row=row_index, column=cost_col_idx, value=base_cost)
        # 代发成本
        sheet.cell(row=row_index, column=dropship_col_idx, value=dropship_cost)
        # 明细总成本=成本+代发成本
        sheet.cell(
            row=row_index,
            column=grand_total_col_idx,
            value=f"=IFERROR({get_column_letter(cost_col_idx)}{row_index}+{get_column_letter(dropship_col_idx)}{row_index},0)",
        )
        # 空列留空
        sheet.cell(row=row_index, column=empty_col_idx, value="")
        # 义乳/义臀成本（空列之后）
        sheet.cell(row=row_index, column=moving_col_idx, value=moving_cost)
        # 枕芯成本
        sheet.cell(row=row_index, column=pillow_col_idx, value=pillow_cost)
        # 硅胶/电动成本
        sheet.cell(row=row_index, column=others_col_idx, value=others_cost)

    # 在明细数据底部添加合计行
    data_start_row = 2
    data_end_row = len(base_costs) + 1
    summary_row = data_end_row + 1

    # 合计行标记在"商家/店铺"列
    if shop_name_col_idx:
        sheet.cell(row=summary_row, column=shop_name_col_idx, value="合计")

    sheet.cell(
        row=summary_row,
        column=cost_col_idx,
        value=f"=SUM({get_column_letter(cost_col_idx)}{data_start_row}:{get_column_letter(cost_col_idx)}{data_end_row})",
    )
    sheet.cell(
        row=summary_row,
        column=moving_col_idx,
        value=f"=SUM({get_column_letter(moving_col_idx)}{data_start_row}:{get_column_letter(moving_col_idx)}{data_end_row})",
    )
    sheet.cell(
        row=summary_row,
        column=dropship_col_idx,
        value=f"=SUM({get_column_letter(dropship_col_idx)}{data_start_row}:{get_column_letter(dropship_col_idx)}{data_end_row})",
    )
    # 明细底部合计：总成本为（成本+代发）逐行公式的求和
    sheet.cell(
        row=summary_row,
        column=grand_total_col_idx,
        value=f"=SUM({get_column_letter(grand_total_col_idx)}{data_start_row}:{get_column_letter(grand_total_col_idx)}{data_end_row})",
    )
    sheet.cell(
        row=summary_row,
        column=pillow_col_idx,
        value=f"=SUM({get_column_letter(pillow_col_idx)}{data_start_row}:{get_column_letter(pillow_col_idx)}{data_end_row})",
    )
    sheet.cell(
        row=summary_row,
        column=others_col_idx,
        value=f"=SUM({get_column_letter(others_col_idx)}{data_start_row}:{get_column_letter(others_col_idx)}{data_end_row})",
    )

    return (
        cost_col_idx,
        dropship_col_idx,
        grand_total_col_idx,
        moving_col_idx,
        pillow_col_idx,
        others_col_idx,
    )


def process_excel_file(
    file_path,
    price_data,
    shop_data,
    moving_costs_data,
    pillow_cost_data,
    others_cost_data,
    output_dir=None,
    overwrite=False,
):
    """处理单个Excel文件，输出到新文件。"""
    print(f"处理文件: {file_path}")

    # 生成输出文件路径（支持自定义目录与覆盖策略）
    output_file_path = build_output_file_path(file_path, output_dir, overwrite)

    try:
        # 对于 .xls 文件，样式可能无法保留
        if file_path.endswith(".xls"):
            print("警告: .xls 文件格式较旧，可能无法保留原始样式。")
            df = pd.read_excel(file_path)
        else:
            # 使用 openpyxl 引擎读取以支持后续的样式保留写入
            df = pd.read_excel(file_path, engine="openpyxl")

        # 检查必要的列是否存在
        required_columns = ["卖家备注", "商家/店铺"]
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            msg = f"文件缺少列 {missing_columns}"
            print(f"警告: {msg}，跳过处理")
            return ProcessResult(
                input_path=file_path,
                success=False,
                status="skipped",
                message=msg,
            )

        # 计算成本、代发成本、义乳/义臀成本、枕芯成本、硅胶/电动成本，同时识别店铺
        base_costs = []  # 成本
        dropship_costs = []  # 代发成本
        moving_costs = []  # 义乳/义臀成本
        pillow_costs = []  # 枕芯成本
        others_costs = []  # 硅胶/电动成本
        grand_total_costs = []  # 总成本
        shop_names = []  # 店铺名称（直接使用"商家/店铺"列去空白）
        overseas_count = 0  # 海外订单计数

        for idx, row in df.iterrows():
            seller_note = str(row["卖家备注"]) if not pd.isna(row["卖家备注"]) else ""
            shop_column = str(row["商家/店铺"]) if not pd.isna(row["商家/店铺"]) else ""

            # 检查是否包含“发海外”关键词（忽略大小写）
            is_overseas = "发海外" in seller_note.lower()

            if is_overseas:
                # 如果是海外订单，所有字段置空
                base_costs.append("")
                dropship_costs.append("")
                moving_costs.append("")
                pillow_costs.append("")
                others_costs.append("")
                grand_total_costs.append("")
                shop_names.append("")
                overseas_count += 1
            else:
                # 成本匹配：仅从卖家备注中匹配,同时计算所有成本类型
                base_cost, dropship_cost, _, moving_cost, pillow_cost, others_cost = (
                    match_price(
                        seller_note,
                        price_data,
                        moving_costs_data,
                        pillow_cost_data,
                        others_cost_data,
                    )
                )

                base_costs.append(base_cost if base_cost is not None else "")
                dropship_costs.append(
                    dropship_cost if dropship_cost is not None else ""
                )

                # 义乳/义臀成本：None表示匹配错误，留空；0表示未匹配
                moving_costs.append(moving_cost if moving_cost is not None else "")

                # 枕芯成本：None表示匹配错误，留空；0表示未匹配
                pillow_costs.append(pillow_cost if pillow_cost is not None else "")

                # 硅胶/电动成本：None或0都可能出现
                others_costs.append(
                    others_cost if (others_cost is not None and others_cost > 0) else 0
                )

                # 计算总成本：成本 + 代发成本
                total_parts = []
                if base_cost is not None:
                    total_parts.append(base_cost)
                if dropship_cost is not None:
                    total_parts.append(dropship_cost)

                grand_total = sum(total_parts) if total_parts else None
                grand_total_costs.append(grand_total if grand_total is not None else "")

                # 店铺名称：直接使用“商家/店铺”原值（去除前后空白）
                shop_name = shop_column.strip()
                shop_names.append(
                    shop_name
                )  # 对于 .xlsx 文件,使用 openpyxl 写入以保留样式
        if file_path.endswith(".xlsx"):
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # 检查工作表是否存在
            if sheet is None:
                msg = "无法获取活动工作表"
                print(f"警告: {msg},跳过处理")
                return ProcessResult(
                    input_path=file_path,
                    success=False,
                    status="failed",
                    message=msg,
                )

            # 检查工作表是否为空
            if sheet.max_row < 1:
                msg = "工作表为空"
                print(f"警告: {msg},跳过处理")
                return ProcessResult(
                    input_path=file_path,
                    success=False,
                    status="failed",
                    message=msg,
                )

            # 将当前sheet重命名为"成本明细"
            sheet.title = "成本明细"

            # 查找或创建表头
            header_row = sheet[1]
            if header_row is None:
                msg = "无法读取表头"
                print(f"警告: {msg},跳过处理")
                return ProcessResult(
                    input_path=file_path,
                    success=False,
                    status="failed",
                    message=msg,
                )

            header = [cell.value for cell in header_row]

            # 使用源数据中的“商家/店铺”列进行统计（不再新增店铺编号/店铺名称列）
            if "商家/店铺" in header:
                shop_name_col_idx = header.index("商家/店铺") + 1
            else:
                print("警告: 缺少‘商家/店铺’列，无法进行店铺统计")
                shop_name_col_idx = None

            # 处理成本明细sheet
            column_indices = process_cost_detail_sheet(
                sheet,
                base_costs,
                dropship_costs,
                moving_costs,
                pillow_costs,
                others_costs,
                shop_names,
                shop_name_col_idx,
            )

            if column_indices is None:
                msg = "成本明细列写入失败"
                return ProcessResult(
                    input_path=file_path,
                    success=False,
                    status="failed",
                    message=msg,
                )

            (
                cost_col_idx,
                dropship_col_idx,
                grand_total_col_idx,
                moving_col_idx,
                pillow_col_idx,
                others_col_idx,
            ) = column_indices

            # 创建独立的店铺统计sheet
            if shop_name_col_idx:
                unique_shop_names = sorted({name for name in shop_names if name})
                data_end_row = len(base_costs) + 1
                process_shop_summary_sheet(
                    workbook,
                    "成本明细",  # 成本明细sheet名称
                    unique_shop_names,
                    2,  # 数据起始行
                    data_end_row,  # 数据结束行
                    get_column_letter(
                        shop_name_col_idx
                    ),  # 店铺名称列字母（源列"商家/店铺"）
                    get_column_letter(cost_col_idx),  # 成本列字母
                    get_column_letter(moving_col_idx),  # 义乳/义臀成本列字母
                    get_column_letter(dropship_col_idx),  # 代发成本列字母
                    get_column_letter(grand_total_col_idx),  # 总成本列字母
                    get_column_letter(pillow_col_idx),  # 枕芯成本列字母
                    get_column_letter(others_col_idx),  # 硅胶/电动成本列字母
                )

            workbook.save(output_file_path)
            print(f"已保存处理后的文件: {output_file_path}")
        else:
            # 对于 .xls 文件，使用 pandas 写入，强制输出为 .xlsx
            # 不再新增店铺编号/店铺名称列，保留原“商家/店铺”列
            df["成本"] = base_costs
            df["代发成本"] = dropship_costs
            df["总成本"] = grand_total_costs
            df["义乳/义臀成本"] = moving_costs
            df["枕芯成本"] = pillow_costs
            df["硅胶/电动成本"] = others_costs
            df.to_excel(output_file_path, index=False, engine="openpyxl")
            print(f"已保存处理后的文件: {output_file_path}")

        # 统计匹配情况
        matched_count = sum(1 for cost in base_costs if cost != "")
        total_count = len(base_costs)
        print(f"匹配成功: {matched_count}/{total_count} 条记录")
        if overseas_count > 0:
            print(f"海外订单: {overseas_count} 条（已置空，需手动补全）")

        return ProcessResult(
            input_path=file_path,
            success=True,
            status="success",
            message="处理成功",
            output_path=output_file_path,
            matched_count=matched_count,
            total_count=total_count,
            overseas_count=overseas_count,
        )

    except Exception as e:
        error_msg = f"处理文件时出错: {e}"
        print(error_msg)
        return ProcessResult(
            input_path=file_path,
            success=False,
            status="failed",
            message=str(e),
        )


def process_single_file(
    file_path,
    json_path,
    shop_json_path,
    moving_costs_json_path,
    pillow_cost_json_path,
    others_json_path,
):
    """处理单个Excel文件"""
    # 加载价格数据
    print("加载价格数据...")
    price_data = load_price_data(json_path)
    print(f"已加载 {len(price_data)} 个类别的价格数据")

    # 加载店铺数据
    shop_data = load_shop_data(shop_json_path)
    if shop_data:
        print(f"已加载 {len(shop_data)} 个店铺数据")

    # 加载动销成本数据
    moving_costs_data = load_moving_costs(moving_costs_json_path)
    if moving_costs_data:
        print(f"已加载 {len(moving_costs_data)} 条动销成本数据")

    # 加载枕芯成本数据
    pillow_cost_data = load_pillow_cost(pillow_cost_json_path)
    if pillow_cost_data:
        print(f"已加载 {len(pillow_cost_data)} 种尺寸的枕芯成本数据")

    # 加载硅胶/电动成本数据
    others_cost_data = load_others_cost(others_json_path)
    if others_cost_data:
        print(f"已加载 {len(others_cost_data)} 条硅胶/电动成本数据")

    # 处理文件
    result = process_excel_file(
        file_path,
        price_data,
        shop_data,
        moving_costs_data,
        pillow_cost_data,
        others_cost_data,
    )
    if result.success:
        print(f"\n处理完成！")
        return True
    else:
        print(f"\n处理失败！原因: {result.message}")
        return False


def _make_input_func(base_path: str) -> Tuple[Callable[[str], str], bool]:
    """创建输入函数。

    优先使用 prompt_toolkit 以支持：
    - Tab 自动补全路径
    - 输入历史记录

    如果 prompt_toolkit 不可用，则回退到内置 input()。

    参数:
        base_path (str): 程序运行目录（脚本目录或 exe 所在目录）

    返回:
        tuple: (input_func, tab_enabled)
    """

    try:
        # 使用动态导入，避免在未安装依赖时触发编辑器的“无法解析导入”提示。
        import importlib

        prompt_toolkit = importlib.import_module("prompt_toolkit")
        completion_mod = importlib.import_module("prompt_toolkit.completion")
        history_mod = importlib.import_module("prompt_toolkit.history")

        PromptSession = getattr(prompt_toolkit, "PromptSession")
        PathCompleter = getattr(completion_mod, "PathCompleter")
        FileHistory = getattr(history_mod, "FileHistory")

        history_file = os.path.join(base_path, ".cost_calculator_history")
        session = PromptSession(
            completer=PathCompleter(expanduser=True),
            history=FileHistory(history_file),
        )

        def _prompt_toolkit_input(prompt_text: str) -> str:
            return session.prompt(prompt_text)

        return _prompt_toolkit_input, True
    except Exception:
        # 任何导入/初始化失败都直接回退，避免影响主流程
        return input, False


def _print_help() -> None:
    """打印交互模式帮助。"""

    print(
        """
可用命令：
  help / ?           显示帮助
  exit / quit / q    退出程序
  cd <目录>          切换当前目录（影响相对路径与 ls）
  ls [目录]          列出目录下的 Excel 文件（.xlsx/.xls）
    outdir             查看当前输出目录策略
    outdir <目录>      设置输出目录（所有结果统一输出到该目录）
    outdir reset       重置为“输出到源文件同目录”
    overwrite on/off   设置同名文件覆盖策略

输入方式：
  - 直接输入 Excel 文件路径（支持拖拽进来）
  - 可一次输入多个路径（用空格分隔；带空格的路径请用引号包住）
  - 输入目录时，会自动处理该目录下的所有 .xlsx/.xls（不递归）
  - 支持通配符：例如 *.xlsx
""".strip()
    )


def _normalize_user_path(raw_path: str, cwd: str) -> str:
    """规范化用户输入路径。

    - 去掉首尾引号
    - 展开环境变量与 ~
    - 相对路径基于 cwd
    """

    cleaned = raw_path.strip().strip("\"").strip("'")
    cleaned = os.path.expandvars(os.path.expanduser(cleaned))
    if not cleaned:
        return ""
    if os.path.isabs(cleaned):
        return os.path.normpath(cleaned)
    return os.path.normpath(os.path.join(cwd, cleaned))


def _expand_cli_tokens(tokens: Sequence[str], cwd: str) -> List[str]:
    """将 token 扩展为实际路径列表。

    支持：
    - 通配符（glob）
    - 相对路径
    """

    results: List[str] = []
    for token in tokens:
        token = token.strip()
        if not token:
            continue

        # 先规范化，再尝试 glob。
        norm = _normalize_user_path(token, cwd)
        if not norm:
            continue

        # glob 需要保留通配符：如果用户输入包含 * ? [，就按 glob 展开。
        if any(ch in token for ch in ("*", "?", "[")):
            matches = glob.glob(norm)
            if matches:
                results.extend([os.path.normpath(p) for p in matches])
            else:
                results.append(norm)
        else:
            results.append(norm)

    # 去重但保序
    seen = set()
    deduped: List[str] = []
    for p in results:
        if p not in seen:
            deduped.append(p)
            seen.add(p)
    return deduped


def _list_excel_files(dir_path: str) -> List[str]:
    """列出目录下的 Excel 文件（不递归）。"""

    try:
        p = Path(dir_path)
        if not p.exists() or not p.is_dir():
            return []
        files = []
        for ext in ("*.xlsx", "*.xls"):
            files.extend([str(x) for x in p.glob(ext)])
        files.sort(key=lambda x: x.lower())
        return files
    except Exception:
        return []


def _parse_user_input_to_paths(user_input: str, cwd: str) -> List[str]:
    """把一行用户输入解析成路径列表。

    说明：
    - 使用 shlex.split 兼容带引号路径
    - 若解析失败，则把整行当作一个路径
    """

    try:
        tokens = shlex.split(user_input, posix=False)
    except Exception:
        tokens = [user_input]
    return _expand_cli_tokens(tokens, cwd)


def main():
    """主函数"""
    # 确定数据文件的基本路径（适用于脚本和打包后的exe）
    if getattr(sys, "frozen", False):
        # 如果作为打包后的应用运行，基本路径是exe文件所在的目录
        base_path = os.path.dirname(sys.executable)
    else:
        # 如果作为脚本运行，基本路径是脚本所在的目录
        base_path = os.path.dirname(os.path.abspath(__file__))

    # JSON文件路径（支持 exe 内嵌资源）
    json_path = resolve_resource_path("size_material_price.json")
    shop_json_path = resolve_resource_path("shop.json")
    moving_costs_json_path = resolve_resource_path("moving_and_selling_costs.json")
    pillow_cost_json_path = resolve_resource_path("pillow_cost.json")
    others_json_path = resolve_resource_path("others.json")

    # 检查必要 JSON 是否存在
    if not os.path.exists(json_path):
        print(f"错误: 找不到价格配置文件 {json_path}")
        return

    # 配置只加载一次（提升多文件处理效率）
    print("加载价格数据...")
    price_data = load_price_data(json_path)
    print(f"已加载 {len(price_data)} 个类别的价格数据")

    # 店铺数据目前不再使用，但保留旧接口
    shop_data = load_shop_data(shop_json_path)

    moving_costs_data = load_moving_costs(moving_costs_json_path)
    if moving_costs_data:
        print(f"已加载 {len(moving_costs_data)} 条动销成本数据")

    pillow_cost_data = load_pillow_cost(pillow_cost_json_path)
    if pillow_cost_data:
        print(f"已加载 {len(pillow_cost_data)} 种尺寸的枕芯成本数据")

    others_cost_data = load_others_cost(others_json_path)
    if others_cost_data:
        print(f"已加载 {len(others_cost_data)} 条硅胶/电动成本数据")

    # 会话日志
    log_path, write_log = create_session_log(base_path)
    if log_path:
        print(f"日志文件: {log_path}")
    write_log("程序启动")

    input_func, tab_enabled = _make_input_func(base_path)
    if tab_enabled:
        print("提示: 已启用 Tab 路径补全（prompt_toolkit）")
    else:
        print("提示: 未启用 Tab 补全（可安装 prompt_toolkit 获得更佳输入体验）")

    # 输出策略
    output_options: Dict[str, object] = {
        "output_dir": None,
        "overwrite": False,
    }
    write_log("输出策略: output_dir=<源文件目录>, overwrite=off")

    # 当前目录：影响相对路径解析与 ls
    cwd = os.getcwd()

    # 先处理命令行参数（可一次传入多个文件）
    initial_args = sys.argv[1:]
    if initial_args:
        initial_paths = _expand_cli_tokens(initial_args, cwd)
        batch_results: List[ProcessResult] = []
        for p in initial_paths:
            _process_targets = [p]
            for target in _process_targets:
                batch_results.extend(
                    _handle_one_target(
                    target,
                    price_data,
                    shop_data,
                    moving_costs_data,
                    pillow_cost_data,
                    others_cost_data,
                    output_options,
                    write_log,
                )
                )
        print_batch_summary(batch_results, write_log)

    print("\n进入交互模式：继续输入文件/目录，或输入 help 查看命令，输入 exit 退出。")

    # 交互式循环处理
    while True:
        try:
            user_input = input_func(f"[{cwd}]> ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n收到退出信号，程序结束。")
            break

        if not user_input:
            continue

        cmd = user_input.strip().lower()
        if cmd in {"exit", "quit", "q"}:
            print("已退出。")
            write_log("收到退出命令")
            break

        if cmd in {"help", "?"}:
            _print_help()
            continue

        if cmd == "cd" or cmd.startswith("cd "):
            parts = user_input.split(maxsplit=1)
            if len(parts) == 1:
                print(f"当前目录: {cwd}")
                continue
            new_dir = _normalize_user_path(parts[1], cwd)
            if not new_dir or not os.path.isdir(new_dir):
                print(f"错误: 目录不存在 {new_dir}")
                continue
            cwd = new_dir
            write_log(f"切换目录: {cwd}")
            continue

        if cmd == "outdir" or cmd.startswith("outdir "):
            parts = user_input.split(maxsplit=1)
            if len(parts) == 1:
                current_outdir = output_options["output_dir"]
                display_outdir = current_outdir if current_outdir else "<源文件目录>"
                overwrite_text = "on" if output_options["overwrite"] else "off"
                print(f"当前输出目录: {display_outdir}")
                print(f"覆盖策略: overwrite {overwrite_text}")
                continue

            arg = parts[1].strip()
            if arg.lower() in {"reset", "clear"}:
                output_options["output_dir"] = None
                print("已重置输出目录为源文件目录。")
                write_log("输出目录重置为源文件目录")
                continue

            target_outdir = _normalize_user_path(arg, cwd)
            if not target_outdir:
                print("错误: 输出目录不能为空")
                continue
            try:
                os.makedirs(target_outdir, exist_ok=True)
            except Exception as e:
                print(f"错误: 无法创建输出目录 {target_outdir}，原因: {e}")
                continue
            output_options["output_dir"] = target_outdir
            print(f"已设置输出目录: {target_outdir}")
            write_log(f"设置输出目录: {target_outdir}")
            continue

        if cmd == "overwrite on" or cmd == "overwrite off":
            is_on = cmd.endswith("on")
            output_options["overwrite"] = is_on
            print(f"已设置覆盖策略: overwrite {'on' if is_on else 'off'}")
            write_log(f"覆盖策略变更: overwrite {'on' if is_on else 'off'}")
            continue

        if cmd == "ls" or cmd.startswith("ls "):
            parts = user_input.split(maxsplit=1)
            target_dir = cwd if len(parts) == 1 else _normalize_user_path(parts[1], cwd)
            excel_files = _list_excel_files(target_dir)
            if not excel_files:
                print("未找到 Excel 文件。")
            else:
                print("Excel 文件列表：")
                for f in excel_files:
                    print(f"  - {f}")
            continue

        # 普通输入：按路径处理（支持多个/通配符/目录）
        targets = _parse_user_input_to_paths(user_input, cwd)
        if not targets:
            print("未解析到有效路径，请重试（输入 help 查看格式）。")
            continue

        batch_results: List[ProcessResult] = []
        for target in targets:
            batch_results.extend(
                _handle_one_target(
                    target,
                    price_data,
                    shop_data,
                    moving_costs_data,
                    pillow_cost_data,
                    others_cost_data,
                    output_options,
                    write_log,
                )
            )

        print_batch_summary(batch_results, write_log)


def _handle_one_target(
    target: str,
    price_data,
    shop_data,
    moving_costs_data,
    pillow_cost_data,
    others_cost_data,
    output_options: Dict[str, object],
    write_log: Callable[[str], None],
) -> List[ProcessResult]:
    """处理一个目标（文件 or 目录）。"""

    results: List[ProcessResult] = []

    if not target:
        return results

    if not os.path.exists(target):
        msg = f"路径不存在: {target}"
        print(f"错误: {msg}")
        results.append(
            ProcessResult(
                input_path=target,
                success=False,
                status="failed",
                message=msg,
            )
        )
        return results

    if os.path.isdir(target):
        excel_files = _list_excel_files(target)
        if not excel_files:
            msg = "目录下未找到 Excel 文件"
            print(f"{msg}: {target}")
            results.append(
                ProcessResult(
                    input_path=target,
                    success=False,
                    status="skipped",
                    message=msg,
                )
            )
            return results
        print(f"\n将处理目录: {target}（共 {len(excel_files)} 个文件）")
        write_log(f"目录批处理: {target} | 文件数: {len(excel_files)}")
        for f in excel_files:
            results.extend(
                _handle_one_target(
                    f,
                    price_data,
                    shop_data,
                    moving_costs_data,
                    pillow_cost_data,
                    others_cost_data,
                    output_options,
                    write_log,
                )
            )
        return results

    # 文件处理
    if not os.path.isfile(target):
        msg = "不是文件"
        print(f"错误: {msg} {target}")
        results.append(
            ProcessResult(
                input_path=target,
                success=False,
                status="failed",
                message=msg,
            )
        )
        return results

    if not (target.lower().endswith(".xlsx") or target.lower().endswith(".xls")):
        msg = "非 Excel 文件"
        print(f"跳过: {msg} {target}")
        results.append(
            ProcessResult(
                input_path=target,
                success=False,
                status="skipped",
                message=msg,
            )
        )
        return results

    print(f"\n将处理文件: {target}")
    write_log(f"开始处理: {target}")

    output_dir = output_options.get("output_dir")
    overwrite = bool(output_options.get("overwrite"))

    result = process_excel_file(
        target,
        price_data,
        shop_data,
        moving_costs_data,
        pillow_cost_data,
        others_cost_data,
        output_dir=output_dir,
        overwrite=overwrite,
    )

    if result.success:
        print("处理完成。")
        write_log(
            f"处理成功: {target} -> {result.output_path} | 匹配 {result.matched_count}/{result.total_count}"
        )
    else:
        print(f"处理失败。原因: {result.message}")
        write_log(f"处理失败: {target} | 原因: {result.message}")

    results.append(result)
    return results


if __name__ == "__main__":
    main()
